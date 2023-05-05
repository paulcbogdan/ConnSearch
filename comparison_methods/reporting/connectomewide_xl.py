import os
from copy import copy
from pathlib import Path
from typing import Union, Optional

import numpy as np
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


def create_xl_table(df_intra, df_inter, fp_xl):
    '''
    This code nicely formats the connectome-wide results table into .xlsx style.
        Honestly, the code is just tedious and not interesting.
    :param df_intra:
    :param df_inter:
    :param fp_xl:
    :return:
    '''
    from openpyxl.cell.cell import MergedCell

    if os.path.isfile(fp_xl): os.remove(fp_xl)
    wb = openpyxl.Workbook()
    del wb['Sheet']
    sheet_name = 'Results'
    ws = wb.create_sheet(sheet_name)

    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(df_intra.columns))
    cell = ws.cell(row=1, column=1)
    cell.value = 'Within Network'
    cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(fp_xl)

    append_df_to_excel(fp_xl, df_intra, index=False, sheet_name=sheet_name)
    wb = openpyxl.load_workbook(fp_xl)
    ws = wb[sheet_name]
    ws.merge_cells(start_row=3 + len(df_intra), start_column=1,
                   end_row=3 + len(df_intra), end_column=len(df_intra.columns))
    cell = ws.cell(row=3 + len(df_intra), column=1)
    cell.value = 'Between Network'
    cell.alignment = Alignment(horizontal='center', vertical='center')

    for column_cells in ws.columns:
        new_column_length = 0
        for cell in column_cells:
            if isinstance(cell, MergedCell):
                continue
            if np.isreal(cell.value):
                cell_len = 5
            else:
                cell_len = len(str(cell.value))
            if cell_len > new_column_length:
                new_column_length = cell_len

        new_column_letter = (get_column_letter(column_cells[0].column))

        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length * 1.23
    wb.save(fp_xl)

    append_df_to_excel(fp_xl, df_inter, index=False, sheet_name=sheet_name)

    wb = openpyxl.load_workbook(fp_xl)  # clear bold created by append_df_to_excel
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            cell.border = None
            cell.font = Font(name='Times New Roman', bold=False)

    # set mini titles to bold
    ws.cell(row=3 + len(df_intra), column=1).font = Font(name='Times New Roman',
                                                         bold=True)
    ws.cell(row=1, column=1).font = Font(name='Times New Roman', bold=True)

    colA = ws['A']
    for cell in colA:
        cell.border = Border(right=Side(style='thin'))

    row_headers = list(ws[2]) + list(ws[4 + len(df_intra)])
    for i, cell in enumerate(row_headers):
        if cell.col_idx == 1:
            cell.border = Border(right=Side(style='thin'),
                                 bottom=Side(style='thin'))
        else:
            cell.border = Border(bottom=Side(style='thin'))

    colE = ws['E']
    for cell in colE:
        if np.isreal(cell.value):
            cell.number_format = '0.000'

    for letter in ['B', 'C', 'D', 'E']:
        for cell in ws[letter]:
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center')

    wb.save(fp_xl)


def copy_excel_cell_range(
        src_ws: openpyxl.worksheet.worksheet.Worksheet,
        min_row: int = None,
        max_row: int = None,
        min_col: int = None,
        max_col: int = None,
        tgt_ws: openpyxl.worksheet.worksheet.Worksheet = None,
        tgt_min_row: int = 1,
        tgt_min_col: int = 1,
        with_style: bool = True
) -> openpyxl.worksheet.worksheet.Worksheet:
    # Code taken from: https://stackoverflow.com/questions/38074678/append-existing-excel-sheet-with-new-dataframe-using-python-pandas/38075046#38075046

    """
    copies all cells from the source worksheet [src_ws] starting from [min_row] row
    and [min_col] column up to [max_row] row and [max_col] column
    to target worksheet [tgt_ws] starting from [tgt_min_row] row
    and [tgt_min_col] column.

    @param src_ws:  source worksheet
    @param min_row: smallest row index in the source worksheet (1-based index)
    @param max_row: largest row index in the source worksheet (1-based index)
    @param min_col: smallest column index in the source worksheet (1-based index)
    @param max_col: largest column index in the source worksheet (1-based index)
    @param tgt_ws:  target worksheet.
                    If None, then the copy will be done to the same (source) worksheet.
    @param tgt_min_row: target row index (1-based index)
    @param tgt_min_col: target column index (1-based index)
    @param with_style:  whether to copy cell style. Default: True

    @return: target worksheet object
    """
    if tgt_ws is None:
        tgt_ws = src_ws

    # https://stackoverflow.com/a/34838233/5741205
    for row in src_ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
        for cell in row:
            tgt_cell = tgt_ws.cell(
                row=cell.row + tgt_min_row - 1,
                column=cell.col_idx + tgt_min_col - 1,
                value=cell.value
            )
            if with_style and cell.has_style:
                # tgt_cell._style = copy(cell._style)
                tgt_cell.font = copy(cell.font)
                tgt_cell.border = copy(cell.border)
                tgt_cell.fill = copy(cell.fill)
                tgt_cell.number_format = copy(cell.number_format)
                tgt_cell.protection = copy(cell.protection)
                tgt_cell.alignment = copy(cell.alignment)
    return tgt_ws


def append_df_to_excel(
        filename: Union[str, Path],
        df: pd.DataFrame,
        sheet_name: str = 'Sheet1',
        startrow: Optional[int] = None,
        max_col_width: int = 30,
        autofilter: bool = False,
        fmt_int: str = "#,##0",
        fmt_float: str = "#,##0.00",
        fmt_date: str = "yyyy-mm-dd",
        fmt_datetime: str = "yyyy-mm-dd hh:mm",
        truncate_sheet: bool = False,
        storage_options: Optional[dict] = None,
        **to_excel_kwargs
) -> None:
    # Code taken from: https://stackoverflow.com/questions/38074678/append-existing-excel-sheet-with-new-dataframe-using-python-pandas/38075046#38075046

    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param max_col_width: maximum column width in Excel. Default: 40
    @param autofilter: boolean - whether add Excel autofilter or not. Default: False
    @param fmt_int: Excel format for integer numbers
    @param fmt_float: Excel format for float numbers
    @param fmt_date: Excel format for dates
    @param fmt_datetime: Excel format for datetime's
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param storage_options: dict, optional
        Extra options that make sense for a particular storage connection, e.g. host, port,
        username, password, etc., if using a URL that will be parsed by fsspec, e.g.,
        starting “s3://”, “gcs://”.
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:


                           fmt_datetime="dd.mm.yyyy hh:mm")

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """

    def set_column_format(ws, column_letter, fmt):
        for cell in ws[column_letter]:
            cell.number_format = fmt

    filename = Path(filename)
    file_exists = filename.is_file()
    # process parameters
    # calculate first column number
    # if the DF will be written using `index=True`, then `first_col = 2`, else `first_col = 1`
    first_col = int(to_excel_kwargs.get("index", True)) + 1
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    # save content of existing sheets
    if file_exists:
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        sheet_exists = sheet_name in sheet_names
        sheets = {ws.title: ws for ws in wb.worksheets}

    with pd.ExcelWriter(
            filename.with_suffix(".xlsx"),
            engine="openpyxl",
            mode="a" if file_exists else "w",
            if_sheet_exists="new" if file_exists else None,
            date_format=fmt_date,
            datetime_format=fmt_datetime,
            storage_options=storage_options
    ) as writer:
        if file_exists:
            # try to open an existing workbook
            writer.book = wb
            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)
            # copy existing sheets
            writer.sheets = sheets
        else:
            # file doesn't exist, we are creating a new one
            startrow = 0

        # write out the DataFrame to an ExcelWriter
        df.to_excel(writer, sheet_name=sheet_name, **to_excel_kwargs)
        worksheet = writer.sheets[sheet_name]

        if autofilter:
            worksheet.auto_filter.ref = worksheet.dimensions

        for xl_col_no, dtyp in enumerate(df.dtypes, first_col):
            col_no = xl_col_no - first_col
            width = max(df.iloc[:, col_no].astype(str).str.len().max(),
                        len(df.columns[col_no]) + 6)
            width = min(max_col_width, width)
            column_letter = get_column_letter(xl_col_no)
            worksheet.column_dimensions[column_letter].width = width
            if np.issubdtype(dtyp, np.integer):
                set_column_format(worksheet, column_letter, fmt_int)
            if np.issubdtype(dtyp, np.floating):
                set_column_format(worksheet, column_letter, fmt_float)

    if file_exists and sheet_exists:
        # move (append) rows from new worksheet to the `sheet_name` worksheet
        wb = load_workbook(filename)
        # retrieve generated worksheet name
        new_sheet_name = set(wb.sheetnames) - set(sheet_names)
        if new_sheet_name:
            new_sheet_name = list(new_sheet_name)[0]
        # copy rows written by `df.to_excel(...)` to
        copy_excel_cell_range(
            src_ws=wb[new_sheet_name],
            tgt_ws=wb[sheet_name],
            tgt_min_row=startrow + 1,
            with_style=True
        )
        # remove new (generated by Pandas) worksheet
        del wb[new_sheet_name]
        wb.save(filename)
        wb.close()
